from pathlib import Path
import tempfile

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from skewb_tnoodle import SkewbTNoodle
from PIL import Image, ImageDraw


# =========================================================
# 1. 你的 move 設定
# =========================================================

MOVES = ["r", "r'", "R", "R'", "b", "b'", "B", "B'"]


def move_axis(move: str) -> str:
    """
    判斷是否同軸。
    r 和 r' 算同一軸。
    R 和 R' 算同一軸。
    """
    return move.replace("'", "")

def inverse_move(move: str) -> str:
    """
    單一步驟取反
    r  -> r'
    r' -> r
    """
    move = move.strip().replace("’", "'")

    if move.endswith("'"):
        return move[:-1]
    else:
        return move + "'"


def inverse_alg(alg: str) -> str:
    """
    整串公式取 inverse
    例如:
    r R' B'  ->  B R r'
    """
    moves = alg.strip().split()
    moves = moves[::-1]  # 順序反過來
    moves = [inverse_move(m) for m in moves]
    return " ".join(moves)


# =========================================================
# 2. 產生 3 步以內所有公式
# =========================================================

def generate_scrambles(max_depth=3):
    scrambles = []

    def dfs(current_moves):
        # 只收 1~max_depth 步
        #if 1 <= len(current_moves) <= max_depth:
        if len(current_moves) == max_depth:
            scrambles.append(" ".join(current_moves))

        if len(current_moves) == max_depth:
            return

        for move in MOVES:
            # 第一步一定是 r 或 r'
            if len(current_moves) == 0:
                if move not in ["r", "r'"]:
                    continue

            # 第二步開始不能跟前一步同軸
            if len(current_moves) > 0:
                prev = current_moves[-1]
                if move_axis(move) == move_axis(prev):
                    continue

            dfs(current_moves + [move])

    dfs([])
    return scrambles


# =========================================================
# 3. 產生 Skewb PNG
# =========================================================


def hex_to_rgb(hex_color: str):
    hex_color = hex_color.lstrip("#")
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))


def render_skewb_png(alg: str, output_png_path: Path):
    skewb = SkewbTNoodle()

    # 這裡改成套 inverse
    inv_alg = inverse_alg(alg)
    skewb.apply_alg(inv_alg)

    width, height = skewb.get_preferred_size()

    # 放大倍率
    scale = 3

    img = Image.new("RGB", (width * scale, height * scale), "white")
    draw = ImageDraw.Draw(img)

    transforms = skewb.get_face_transforms()
    base_paths = skewb.get_face_paths()
    scheme = [skewb.color_scheme[ch] for ch in skewb.FACE_ORDER]

    for face in range(6):
        trans = transforms[face]

        for sticker in range(5):
            pts = skewb.apply_transform(base_paths[sticker], trans)
            pts_scaled = [(x * scale, y * scale) for x, y in pts]

            fill_color = hex_to_rgb(scheme[skewb.image[face][sticker]])

            draw.polygon(
                pts_scaled,
                fill=fill_color,
                outline=(0, 0, 0)
            )

    img.save(output_png_path)


# =========================================================
# 4. 生成 Excel
# =========================================================
def create_excel(scrambles, output_xlsx="skewb_4step_scrambles.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Skewb {wb}-step"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, size=11)

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 每 12 個換一組
    ITEMS_PER_ROW = 72

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)

        for idx, alg in enumerate(scrambles):
            group = idx // ITEMS_PER_ROW      # 第幾組，0~5
            pos = idx % ITEMS_PER_ROW         # 該組中的第幾個，0~11

            col = pos + 1
            alg_row = group * 2 + 1           # 1, 3, 5, 7, 9, 11
            img_row = group * 2 + 2           # 2, 4, 6, 8, 10, 12

            col_letter = get_column_letter(col)

            # 設定欄寬
            ws.column_dimensions[col_letter].width = 18

            # 公式列高度
            ws.row_dimensions[alg_row].height = 28

            # 圖片列高度
            ws.row_dimensions[img_row].height = 80

            # 寫入公式
            cell = ws.cell(row=alg_row, column=col)
            cell.value = alg
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

            # 產生圖片
            png_path = tmpdir / f"skewb_{idx + 1:03d}.png"
            render_skewb_png(alg, png_path)

            img = ExcelImage(str(png_path))
            img.width = 125
            img.height = 105

            ws.add_image(img, f"{col_letter}{img_row}")

            img_cell = ws.cell(row=img_row, column=col)
            img_cell.border = border
            img_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 覆蓋舊檔
        output_path = Path(output_xlsx)

        if output_path.exists():
            try:
                output_path.unlink()
            except PermissionError:
                raise PermissionError(
                    f"無法覆蓋 {output_xlsx}，請先關掉 Excel 檔案。"
                )

        wb.save(output_path)

    print(f"完成，共產生 {len(scrambles)} 組")
    print(f"Excel 已輸出：{output_xlsx}")


# =========================================================
# 5. 主程式
# =========================================================

if __name__ == "__main__":
    scrambles = generate_scrambles(max_depth=5)

    print("總數：", len(scrambles))
    print(scrambles[:20])

    create_excel(
        scrambles=scrambles,
        output_xlsx="skewb_5step_layer.xlsx"
    )